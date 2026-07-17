#!/usr/bin/env python3
"""build_excel.py — sinh dist/aiop-catalog.xlsx từ model (_index).

Sheets:
  Modules       — id, dir, class, order, band, tier, axis, backbone, spine, capabilities (tóm tắt)
  Parts         — id, home, owner, layer, at, note
  ConfigParts   — id, home, owner, vi, en
  Links         — from, uses, why
  Foundations   — id, provides, anchor
  Safety        — id, vi, en, anchors
"""
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    print("[excel] WARN: openpyxl chưa cài — pip install openpyxl. Bỏ qua build Excel.")
    sys.exit(0)

REPO = Path(__file__).resolve().parents[6]
DOCSLIB = REPO / "harness" / "plugins" / "hs" / "skills" / "_docslib"
sys.path.insert(0, str(DOCSLIB))

from docslib import load_model  # noqa: E402

HEADER_FONT = Font(bold=True)


def _header(ws, cols: list[str]):
    ws.append(cols)
    for cell in ws[1]:
        cell.font = HEADER_FONT


def _cap_summary(caps: dict) -> str:
    """Tóm tắt capabilities dict thành chuỗi ngắn (các key=True)."""
    if not caps:
        return ""
    return ", ".join(k for k, v in caps.items() if v)


def main():
    dist = REPO / "dist"
    dist.mkdir(parents=True, exist_ok=True)
    out_path = dist / "aiop-catalog.xlsx"

    print("[excel] Load model...")
    model = load_model(REPO / "docs")

    wb = openpyxl.Workbook()
    # Xoá sheet mặc định
    wb.remove(wb.active)

    # ------------------------------------------------------------------
    # Sheet: Modules
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Modules")
    _header(ws, ["id", "dir", "class", "order", "band", "tier", "axis",
                 "backbone", "spine", "capabilities"])
    for m in model.modules:
        ws.append([
            m.id, m.dir, m.klass, m.order, m.band or "",
            m.tier or "", m.axis or "",
            "yes" if m.backbone else "", "yes" if m.spine else "",
            _cap_summary(m.capabilities),
        ])

    # ------------------------------------------------------------------
    # Sheet: Parts
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Parts")
    _header(ws, ["id", "home", "owner", "layer", "at", "note"])
    for pid, p in sorted(model.parts.items()):
        ws.append([
            pid,
            p.get("home", ""),
            p.get("owner", ""),
            p.get("layer", ""),
            p.get("at", ""),
            p.get("note", ""),
        ])

    # ------------------------------------------------------------------
    # Sheet: ConfigParts
    # ------------------------------------------------------------------
    ws = wb.create_sheet("ConfigParts")
    _header(ws, ["id", "home", "owner", "vi", "en"])
    for cid, c in sorted(model.config_parts.items()):
        ws.append([
            cid,
            c.get("home", ""),
            c.get("owner", ""),
            c.get("vi", ""),
            c.get("en", ""),
        ])

    # ------------------------------------------------------------------
    # Sheet: Links
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Links")
    _header(ws, ["from", "uses", "why"])
    for ln in model.links:
        ws.append([
            ln.get("from", ""),
            ln.get("uses", ""),
            ln.get("why", ""),
        ])

    # ------------------------------------------------------------------
    # Sheet: Foundations
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Foundations")
    _header(ws, ["id", "provides", "anchor"])
    for f in model.foundations:
        ws.append([
            f.get("id", ""),
            f.get("provides", ""),
            f.get("anchor", ""),
        ])

    # ------------------------------------------------------------------
    # Sheet: Safety
    # ------------------------------------------------------------------
    ws = wb.create_sheet("Safety")
    _header(ws, ["id", "vi", "en", "anchors"])
    for sfy in model.safety:
        ws.append([
            sfy.get("id", ""),
            sfy.get("vi", ""),
            sfy.get("en", ""),
            sfy.get("anchors", ""),
        ])

    wb.save(out_path)
    print(f"[excel] {len(wb.sheetnames)} sheets → {out_path} ({out_path.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
