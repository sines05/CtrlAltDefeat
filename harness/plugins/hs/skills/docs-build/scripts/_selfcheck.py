#!/usr/bin/env python3
"""_selfcheck.py — kiểm tra artifact sau build_all.py.

In PASS/FAIL cho từng assert theo yêu cầu P4.
Exit 1 nếu có bất kỳ FAIL nào.
"""
import sys
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

REPO = Path(__file__).resolve().parents[6]


def check(label: str, ok: bool, detail: str = "") -> bool:
    status = "PASS" if ok else "FAIL"
    msg = f"[{status}] {label}"
    if detail:
        msg += f" — {detail}"
    print(msg)
    return ok


def main():
    results = []

    # 1. public/index.html tồn tại + >1KB
    idx = REPO / "public" / "index.html"
    ok1 = idx.is_file() and idx.stat().st_size > 1024
    results.append(check("public/index.html tồn tại + >1KB",
                         ok1, f"{idx.stat().st_size if idx.is_file() else 0} bytes"))

    # 2. public/pages/ có ≥10 file
    pages_dir = REPO / "public" / "pages"
    page_files = list(pages_dir.glob("*.html")) if pages_dir.is_dir() else []
    ok2 = len(page_files) >= 10
    results.append(check("public/pages/ có ≥10 file", ok2, f"{len(page_files)} file"))

    # 3. docs/showcase/assets/js/module-m4-data.js tồn tại (nguồn _index vừa sinh)
    #    Lưu ý: build.py BUNDLE tất cả JS_PARTS (bao gồm module-m4-data.js) vào showcase.js —
    #    không copy riêng js/ sang public/. Kiểm tra file nguồn trong showcase/assets/js/.
    m4_src = REPO / "docs" / "showcase" / "assets" / "js" / "module-m4-data.js"
    ok3 = m4_src.is_file()
    results.append(check("docs/showcase/assets/js/module-m4-data.js tồn tại (nguồn _index)", ok3,
                         f"{m4_src.stat().st_size if ok3 else 0} bytes"))

    # 4. showcase.js trong public/ chứa var MOD_M4 (data đã bundle)
    showcase_js = REPO / "public" / "assets" / "showcase.js"
    if showcase_js.is_file():
        js_text = showcase_js.read_text(encoding="utf-8")
        ok4 = "var MOD_M4 =" in js_text
        results.append(check("public/assets/showcase.js chứa var MOD_M4 (data bundle từ _index)", ok4))
    else:
        results.append(check("public/assets/showcase.js tồn tại", False))

    # 5. dist/aiop-docs.md >200 dòng + chứa "PTNT"
    flat_md = REPO / "dist" / "aiop-docs.md"
    if flat_md.is_file():
        content = flat_md.read_text(encoding="utf-8")
        line_count = content.count("\n")
        has_ptnt = "PTNT" in content
        ok5a = line_count > 200
        ok5b = has_ptnt
        results.append(check("dist/aiop-docs.md >200 dòng", ok5a, f"{line_count} dòng"))
        results.append(check("dist/aiop-docs.md chứa 'PTNT'", ok5b))
    else:
        results.append(check("dist/aiop-docs.md tồn tại", False))

    # 6. dist/aiop-catalog.xlsx tồn tại + mở được + 6 sheet
    xlsx = REPO / "dist" / "aiop-catalog.xlsx"
    if xlsx.is_file() and HAS_OPENPYXL:
        try:
            wb = openpyxl.load_workbook(xlsx, read_only=True)
            sheet_count = len(wb.sheetnames)
            wb.close()
            ok6 = sheet_count == 6
            results.append(check("dist/aiop-catalog.xlsx mở được + 6 sheet", ok6,
                                 f"sheets={sheet_count}: {wb.sheetnames if not ok6 else ''}"))
        except Exception as e:
            results.append(check("dist/aiop-catalog.xlsx mở được", False, str(e)))
    elif not HAS_OPENPYXL:
        results.append(check("dist/aiop-catalog.xlsx (openpyxl unavailable)", False,
                             "pip install openpyxl"))
    else:
        results.append(check("dist/aiop-catalog.xlsx tồn tại", False))

    # 7. dist/aiop-docs.pdf tồn tại (nếu lib OK)
    pdf = REPO / "dist" / "aiop-docs.pdf"
    if pdf.is_file():
        results.append(check("dist/aiop-docs.pdf tồn tại", True,
                             f"{pdf.stat().st_size:,} bytes"))
    else:
        print("[INFO] dist/aiop-docs.pdf không tồn tại — có thể xhtml2pdf chưa cài.")

    # Kết quả
    fails = [r for r in results if not r]
    print(f"\n{'='*50}")
    print(f"Tổng: {len(results)} check — PASS: {len(results)-len(fails)}, FAIL: {len(fails)}")
    if fails:
        sys.exit(1)


if __name__ == "__main__":
    main()
